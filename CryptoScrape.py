import openpyxl
from openpyxl import load_workbook, cell
import selenium.webdriver as webdriver
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys

def main(currency):

    """I set it up to the user agent i am using and set up the driver path of the Firefox"""
    user_agent = "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:97.0) Gecko/20100101 Firefox/97.0"
    path = Service("C:\Code\geckodriver.exe")
    FireFox_driver = webdriver.Firefox(service=path)

    """Pass in the url to open and wait for the website to open up fully"""
    url = "https://www.tradingview.com/markets/cryptocurrencies/prices-all/"
    FireFox_driver.get(url)
    FireFox_driver.implicitly_wait(2)

    """I get the value of which currency the website is on as it loads and changes it to BTC by 
    clicking it after it loads."""
    active_currency = FireFox_driver.find_element(By.XPATH, '//span[contains(@class, "modeTitleActive-3iGYBWzh")]')
    if active_currency.text == currency:
            pass
    else:
        FireFox_driver.find_element(By.XPATH, '//input[@type="checkbox"]').click()

    """I load more of the crypto by finding the load more button and execute the click using javascript."""
    for i in range(0,7):
        load_button = FireFox_driver.find_element(By.CLASS_NAME, "tv-load-more__btn")
        FireFox_driver.execute_script("arguments[0].click();", load_button);
        FireFox_driver.implicitly_wait(1)

    """I find the values of all the different elements i want to add to the excel and save it to the appropriate variables."""
    crypto_CATEG = FireFox_driver.find_elements(By.CSS_SELECTOR, '.tv-screener-table__head-left--title-three-lines')
    crypto_NAME = FireFox_driver.find_elements(By.XPATH, '//a[@class="tv-screener__symbol"]')
    crypto_MKTCAP = FireFox_driver.find_elements(By.CSS_SELECTOR, '.tv-screener-table__cell--left+ .tv-screener-table__cell--big')
    crypto_FD_MKTCAP = FireFox_driver.find_elements(By.CSS_SELECTOR,'.tv-screener-table__cell--big:nth-child(3)')
    crypto_LAST = FireFox_driver.find_elements(By.CSS_SELECTOR, '.tv-screener-table__cell--big:nth-child(4)')
    crypto_AVCOINS = FireFox_driver.find_elements(By.CSS_SELECTOR, '.tv-screener-table__cell--big:nth-child(5)')
    crypto_TOTCOINS = FireFox_driver.find_elements(By.CSS_SELECTOR, '.tv-screener-table__cell--big:nth-child(6  )')
    crypto_TRADVOL = FireFox_driver.find_elements(By.CSS_SELECTOR, '.tv-screener-table__cell--big:nth-child(7)')
    crypto_CHG = FireFox_driver.find_elements(By.CSS_SELECTOR, '.tv-screener-table__cell--big:nth-child(8)')

    """I get the number of elements present and loop through that much. The website does vary crypto shown thats
    why i used this method rather than hard coding it in"""
    crypto_NomElement = FireFox_driver.find_element(By.CSS_SELECTOR, '.tv-screener-table__field-value--total')
    crypto_Nom = crypto_NomElement.text
    crypto_Nom = crypto_Nom.split()

    """Open up excel and create a sheet as shown."""
    wb = openpyxl.Workbook()
    sheet = wb.create_sheet("Crypto_Price", 1)

    print(int(crypto_Nom[0]))

    """The code below takes the amount of crypto names from the website and uses it as len. Then for each section
    in rows add each of the different values of the crypto such as marketcap, totalcoins etc
    this is done by looping through another list to fill in the rows horizontalily."""
    sections = ["A","B","C","D","E","F","G","H"]
    i = -1
    for j in range(1,int(crypto_Nom[0])):
        k = 0
        i = i+1
        crypto_values = [crypto_NAME[i],crypto_MKTCAP[i],crypto_FD_MKTCAP[i],crypto_LAST[i],crypto_AVCOINS[i],crypto_TOTCOINS[i],crypto_TRADVOL[i],crypto_CHG[i]]
        for s in sections:
            sheet[f'{s}{j}'] = crypto_values[k].text
            k +=1

    """After the values have been added i save the file and close the broswere"""
    wb.save("CryptoPrice.xlsx")
    FireFox_driver.close()

"""Main to run the code and select between BTC or USD"""
if __name__ == '__main__':
    currency = 'BTC'
    main(currency)
